import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

st.set_page_config(page_title="WMS Performance Report", layout="wide")


st.title("📦 WMS Performance Report")

uploaded_file = st.file_uploader("Upload WMS Report", type=["xlsx"])

if uploaded_file:
    wb = load_workbook(uploaded_file, data_only=True)
    sheet = wb['Sheet2']
    
    # Get Picker table (MV38:ND47)
    mv_col = column_index_from_string('MV')
    nd_col = column_index_from_string('ND')
    
    # Headers
    headers = []
    for col in range(mv_col, nd_col + 1):
        val = sheet.cell(row=38, column=col).value
        headers.append(val if val else '')
    
    # Data rows
    data = []
    for row in range(39, 48):
        row_data = []
        for col in range(mv_col, nd_col + 1):
            val = sheet.cell(row=row, column=col).value
            row_data.append(val if val is not None else '')
        if row_data[0]:  # Only add if there's a picker name
            data.append(row_data)
    
    df = pd.DataFrame(data, columns=headers)
    
    # Get statistics
    date_val = sheet.cell(row=60, column=367).value
    total_picking_time = sheet.cell(row=63, column=361).value
    total_requests = sheet.cell(row=63, column=362).value
    avg_requests_min = sheet.cell(row=63, column=363).value
    total_kg = sheet.cell(row=63, column=364).value
    total_l = sheet.cell(row=63, column=365).value
    avg_kg_min = sheet.cell(row=63, column=366).value
    avg_l_min = sheet.cell(row=63, column=367).value
    avg_per_min = sheet.cell(row=63, column=368).value
    picking_finish = sheet.cell(row=67, column=362).value
    
    # Find max values for progress bars
    max_time = df['Picking Time'].apply(lambda x: sum(int(i) * 60**(2-idx) for idx, i in enumerate(str(x).split(':'))) if x else 0).max()
    max_requests = df['Requests fulfilled'].apply(lambda x: float(x) if x else 0).max()
    max_kg = df['Kilograms'].apply(lambda x: float(x) if x else 0).max()
    max_l = df['Liters'].apply(lambda x: float(x) if x else 0).max()
    
    # Color function for Avg per min
    def get_avg_color(val):
        try:
            v = float(val)
            if v >= 10:
                return '#90EE90'  # Green
            elif v >= 7:
                return '#FFFF00'  # Yellow
            elif v >= 5:
                return '#FFA500'  # Orange
            else:
                return '#FF6B6B'  # Red
        except:
            return '#FFFFFF'
    
    # Build HTML table
    html = '''
    <style>
        .wms-table {
            border-collapse: collapse;
            width: 100%;
            font-family: Arial, sans-serif;
            font-size: 14px;
        }
        .wms-table th {
            background-color: #4472C4;
            color: white;
            padding: 10px;
            text-align: center;
            border: 1px solid #2F5496;
        }
        .wms-table td {
            padding: 8px;
            border: 1px solid #B4C6E7;
            text-align: center;
        }
        .wms-table tr:nth-child(odd) {
            background-color: #D6DCE4;
        }
        .wms-table tr:nth-child(even) {
            background-color: #EDEDED;
        }
        .picker-name {
            background-color: #C65B5B !important;
            color: white;
            font-weight: bold;
            text-align: left !important;
        }
        .progress-cell {
            position: relative;
            padding: 0 !important;
        }
        .progress-bar {
            height: 100%;
            position: absolute;
            left: 0;
            top: 0;
        }
        .progress-text {
            position: relative;
            z-index: 1;
            padding: 8px;
        }
        .stats-table {
            border-collapse: collapse;
            margin-top: 30px;
            font-family: Arial, sans-serif;
        }
        .stats-table th {
            background-color: #4472C4;
            color: white;
            padding: 10px;
            border: 1px solid #2F5496;
        }
        .stats-table td {
            padding: 10px;
            border: 1px solid #B4C6E7;
            background-color: #D6DCE4;
            text-align: center;
        }
        .stats-title {
            font-size: 18px;
            text-decoration: underline;
            margin-bottom: 10px;
        }
        .date-box {
            background-color: #FFFF00;
            padding: 5px 15px;
            border: 1px solid #000;
            display: inline-block;
        }
    </style>
    '''
    
    html += '<table class="wms-table">'
    
    # Header row
    html += '<tr>'
    for h in headers:
        html += f'<th>{h}</th>'
    html += '</tr>'
    
    # Data rows
    for idx, row in df.iterrows():
        html += '<tr>'
        for col_idx, (col_name, val) in enumerate(row.items()):
            if col_name == 'Picker':
                html += f'<td class="picker-name">{val}</td>'
            elif col_name == 'Picking Time':
                # Blue progress bar
                try:
                    time_parts = str(val).split(':')
                    time_secs = sum(int(i) * 60**(2-idx) for idx, i in enumerate(time_parts))
                    pct = (time_secs / max_time * 100) if max_time > 0 else 0
                except:
                    pct = 0
                html += f'''<td class="progress-cell">
                    <div class="progress-bar" style="width: {pct}%; background-color: #C65B5B;"></div>
                    <div class="progress-text">{val}</div>
                </td>'''
            elif col_name == 'Requests fulfilled':
                try:
                    pct = (float(val) / max_requests * 100) if max_requests > 0 else 0
                except:
                    pct = 0
                html += f'''<td class="progress-cell">
                    <div class="progress-bar" style="width: {pct}%; background-color: #5B9BD5;"></div>
                    <div class="progress-text">{int(float(val)) if val else ""}</div>
                </td>'''
            elif col_name == 'Requests per minute':
                try:
                    html += f'<td>{float(val):.2f}</td>'
                except:
                    html += f'<td>{val}</td>'
            elif col_name == 'Kilograms':
                try:
                    pct = (float(val) / max_kg * 100) if max_kg > 0 else 0
                except:
                    pct = 0
                html += f'''<td class="progress-cell">
                    <div class="progress-bar" style="width: {pct}%; background-color: #FFC000;"></div>
                    <div class="progress-text">{float(val):.2f}</div>
                </td>'''
            elif col_name == 'Liters':
                try:
                    pct = (float(val) / max_l * 100) if max_l > 0 else 0
                except:
                    pct = 0
                html += f'''<td class="progress-cell">
                    <div class="progress-bar" style="width: {pct}%; background-color: #70AD47;"></div>
                    <div class="progress-text">{float(val):.2f}</div>
                </td>'''
            elif col_name in ['Kg per min', 'L per min']:
                try:
                    html += f'<td>{float(val):.2f}</td>'
                except:
                    html += f'<td>{val}</td>'
            elif col_name == 'Avg per min':
                color = get_avg_color(val)
                try:
                    html += f'<td style="background-color: {color}; font-weight: bold;">{float(val):.3f}</td>'
                except:
                    html += f'<td style="background-color: {color};">{val}</td>'
            else:
                html += f'<td>{val}</td>'
        html += '</tr>'
    
    html += '</table>'
    
    # Statistics section
    html += f'''
    <div style="margin-top: 40px;">
        <span class="stats-title">Statistics for {date_val}</span>
        <span class="date-box" style="margin-left: 300px;">{date_val}</span>
    </div>
    <table class="stats-table" style="margin-top: 15px;">
        <tr>
            <th>Total Picking Time</th>
            <th>Total Requests</th>
            <th>Avg Requests per minute</th>
            <th>Total Kg</th>
            <th>Total L</th>
            <th colspan="3">Avg Per minute</th>
        </tr>
        <tr>
            <td>{total_picking_time}</td>
            <td>{total_requests}</td>
            <td>{float(avg_requests_min):.2f if avg_requests_min else ""}</td>
            <td>{total_kg}</td>
            <td>{total_l}</td>
            <td>{avg_kg_min}</td>
            <td>{avg_l_min}</td>
            <td>{float(avg_per_min):.2f if avg_per_min else ""}</td>
        </tr>
    </table>
    <div style="margin-top: 30px;">
        <span class="date-box">{date_val}</span>
    </div>
    <table class="stats-table" style="margin-top: 10px;">
        <tr>
            <th>Picking Finish</th>
            <td>{picking_finish}</td>
        </tr>
    </table>
    '''
    
    st.markdown(html, unsafe_allow_html=True)

else:
    st.info("👆 Upload your WMS report file")
